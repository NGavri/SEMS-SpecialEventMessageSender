#required imports
from kivy.lang import Builder
from kivy.app import App
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.core.window import Window
from kivy.uix.textinput import TextInput
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.clock import Clock
import pywhatkit
from openpyxl import load_workbook
from time import strftime, sleep
from kivy.uix.label import Label
from datetime import datetime
from kivy.uix.popup import Popup
from kivy.graphics import Color,RoundedRectangle,Rectangle
from kivy.uix.image import Image
from kivy.animation import Animation



# Load workbook and set active sheet
wb = load_workbook('dataM.xlsx')
ws = wb.active
ws.title = 'Details'

# Get current month and day

month = strftime("%b")
day = strftime("%d")
da = day + month

kv_string = '''
#: import SlideTransition kivy.uix.screenmanager.SlideTransition
<MyBoxLayout>:
    orientation: 'vertical'
    spacing: 25
    padding: 10

    

    canvas.before:

        Color:
            rgba: 231/255, 231/255, 231/255, 1
        Rectangle:
            pos: self.pos
            size: self.size
    
    Label
        text: "SEMS"
        font_name: 'Inter-Bold.otf'
        font_size: 50
        color: 0,0,0,1


    Label:

    Label:

    Label:

    Label:

    RoundedButton:
        size_hint: None, None
        size: 70,70
        pos: 100,10
        on_press:app.root.transition = SlideTransition(direction = 'left'); app.root.current = 'add_person'
        Image:
            source: 'addUser.png'  
            center_x: self.parent.center_x
            center_y: self.parent.center_y

    RoundedButton:
        size_hint: None, None
        size: 70, 70
        pos: 100,10
        on_press: app.search_people()
        Image:
            source: 'viewList.png'  
            center_x: self.parent.center_x
            center_y: self.parent.center_y







<AddPersonScreen>:
    canvas.before:
        Color:
            rgba: 231/255, 231/255, 231/255, 1
        Rectangle:
            pos: self.pos
            size: self.size

    orientation: 'vertical'
    spacing: 20
    padding: 20

    RoundedButton:
        size_hint: None, None
        size: 50,50
        on_press: 
            app.root.transition = SlideTransition(direction = 'right')
            app.root.current = 'my_box_layout'
        Image:
            source: 'home.png'  
            center_x: self.parent.center_x
            center_y: self.parent.center_y
    Label:
        text: 'Add Person'
        font_name: 'Inter-Bold.otf'
        color: 0,0,0,1
        font_size: 30
        size_hint_y: None
        height: self.texture_size[1]

    RoundedInputBox:
        size_hint: 1,.2
        id: name_input
        hint_text: 'Enter name'
        font_name: 'Roboto-Light.ttf'
        multiline: False

    RoundedInputBox:
        size_hint: 1,.2
        id: event_input
        hint_text: 'Enter event date (Ex: 30Apr)'
        font_name: 'Roboto-Light.ttf'
        multiline: False


    RoundedInputBox:
        size_hint: 1,.2
        id: whatsapp_input
        hint_text: 'Enter whatsapp number (Ex: +1XXXXXXXXXX)'
        font_name: 'Roboto-Light.ttf'
        multiline: False


    RoundedInputBox:
        size_hint: 1,.2
        id: message_input
        hint_text: 'Enter Message you want to send'
        font_name: 'Roboto-Light.ttf'
        multiline: True

    RoundedInputBox:
        size_hint: 1,.2
        id: eventName_input
        hint_text: 'Enter Event Name'
        font_name: 'Roboto-Light.ttf'
        multiline: True


    RoundedButton:
        on_press: 
            if len(name_input.text) > 0 and len(event_input.text) > 0 and len(whatsapp_input.text) > 0 and len(message_input.text) > 0 and len(eventName_input.text) > 0 :app.add_person(name_input.text, event_input.text, whatsapp_input.text, message_input.text, eventName_input.text); name_input.text = ""; event_input.text = ""; whatsapp_input.text = ""; message_input.text = ""; eventName_input.text = ""; app.root.transition = SlideTransition(direction = 'right'); app.root.current = 'my_box_layout'; app.show_popup_message('NOTIFICATION','Person Added Successful')
            else: app.show_popup_message('WARNING','Fill all fields')
        size_hint: None, None
        size: 70, 70
        pos_hint: {'x': 0.0, 'y': 1.0}
        Image:
            source: 'add.png'  
            center_x: self.parent.center_x
            center_y: self.parent.center_y
    



            
<details>
    canvas.before:
        Color:
            rgba: 231/255, 231/255, 231/255, 1
        Rectangle:
            pos: self.pos
            size: self.size

    BoxLayout:
        orientation: 'vertical'
        spacing: 10
        padding: 2


<RoundedButton@Button>
    background_color: 0,0,0,0
    background_normal: ''
    canvas.before:
        Color:
            rgba: 44/255, 43/255, 43/255, 0.27
        RoundedRectangle:
            size: self.size
            pos: self.pos
            radius: [20]
<RoundedInputBox@TextInput>
    background_color: 0,0,0,0
    background_normal: ''
    font_name: 'Roboto-Light.ttf'
    canvas.before:
        Color:
            rgba: 83/255, 142/255, 193/255, 0.45
        RoundedRectangle:
            size: self.size
            pos: self.pos
            radius: [10]


'''

Builder.load_string(kv_string)


class MyBoxLayout(BoxLayout):
    pass

class AddPersonScreen(BoxLayout):
    pass

class LoadingScreen(BoxLayout):
    def on_parent(self, widget, parent):
        if parent:
            # Set the background color to white
            with self.canvas.before:
                Color(231/255, 231/255, 231/255, 1)  # White color
                self.rect = Rectangle(pos=self.pos, size=self.size)
                self.bind(pos=self.update_rect, size=self.update_rect)

            # Add logo image to the screen and animate its opacity
            logo = Image(source='logo.png', opacity=1)
            self.add_widget(logo)
            anim = Animation(opacity=0, duration=2.5)
            anim.start(logo)

    def update_rect(self, *args):
        # Update the position and size of the background rectangle
        self.rect.pos = self.pos
        self.rect.size = self.size

class CurvedPopup(Popup):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.background_color = (0, 0, 0, 0)
        self.separator_height = 0
        self.size_hint = (None, None)
        self.size = (300, 200)

        with self.canvas.before:
            # Draw a rounded rectangle as the background
            Color(0.8, 0.8, 0.8, 1)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[10, 10, 10, 10])

    def on_pos(self, instance, value):
        # Update the position and size of the rounded rectangle when the popup position changes
        self.canvas.before.clear()
        with self.canvas.before:
            Color(0.8, 0.8, 0.8, 1)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[10, 10, 10, 10])

class Search(Popup):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        self.background_color = (0, 0, 0, 0)
        self.separator_height = 0
        self.size_hint = (None, None)
        self.size = (300, 200)
        self.input_text = TextInput(hint_text='Name', multiline=False, background_color=(0, 0, 0, .2))
        self.input_text2 = TextInput(hint_text='Event Name', multiline=False, background_color=(0, 0, 0, .2))
        self.input_text2.bind(on_text_validate=self.findPeople)  # Bind on_text_validate event

        self.layout.add_widget(self.input_text)
        self.layout.add_widget(self.input_text2)
        self.content = self.layout
        with self.canvas.before:
            Color(0.8, 0.8, 0.8, 1)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[10, 10, 10, 10])

    def on_pos(self, instance, value):
        # Update the position and size of the rounded rectangle when the popup position changes
        self.canvas.before.clear()
        with self.canvas.before:
            Color(0.8, 0.8, 0.8, 1)
            RoundedRectangle(pos=self.pos, size=self.size, radius=[10, 10, 10, 10])

    def findPeople(self, instance):
        def show_popup(head, message):
            popup = CurvedPopup(title=head, content=Label(text=message), size_hint=(None, None), size=(400, 200), separator_height=0, background_color=(1, 1, 1, 1))
            popup.open()

        # Read the log file
        log = open("codeLog.txt", "r")
        i = log.readline()
        j = int(i)
        nm = self.input_text.text
        em = self.input_text2.text
        na = str(nm)
        ea = str(em)
        c = 0
        date = ""
        number = ""
        msg = ""
        event = ""

        while j > 0:
            name = ws['A' + str(j)].value
            event = ws['E' + str(j)].value
            ng = str(name)
            eg = str(event)
            if ng == na and eg == ea:
                d = ws['B' + str(j)].value
                num = ws['C' + str(j)].value
                m = ws['D' + str(j)].value
                e = ws['E' + str(j)].value
                date = str(d)
                number = str(num)
                msg = str(m)
                event = str(e)

                c += 1
            j -= 1
        a = str(c)
        detail = "Name: " + ng + '\n' + "Date: " + date + '\n' + "Number: " + number + '\n' + "Message: " + msg + '\n' + "Event: " + event + '\n' + "Number of repetition of same detail: " + a
        if c >= 1:
            show_popup(head="Details", message=detail)
        else:
            show_popup(head="ERROR", message="Person Not Found")

class MyApp(App):
    title = "SEMS"
    icon = "logo.png"

    def build(self):
        Window.size = (360, 640)

        # Create the screen manager
        screen_manager = ScreenManager()

        # Create the loading screen
        loading_screen = LoadingScreen()
        screen = Screen(name='loading')
        screen.add_widget(loading_screen)
        screen_manager.add_widget(screen)

        # Create the add person screen
        add_person_screen = AddPersonScreen()
        screen = Screen(name='add_person')
        screen.add_widget(add_person_screen)
        screen_manager.add_widget(screen)

        # Create the main screen
        my_box_layout = MyBoxLayout()
        screen = Screen(name='my_box_layout')
        screen.add_widget(my_box_layout)
        screen_manager.add_widget(screen)

        # Set the default screen to the loading screen
        screen_manager.current = 'loading'

        # Schedule a callback to switch to the home screen after 5 seconds
        Clock.schedule_once(lambda dt: screen_manager.switch_to(screen_manager.get_screen('my_box_layout')), 5)
        return screen_manager

    def send_whatsapp_message(self, number, message, hour, minute):
        pywhatkit.sendwhatmsg(number, message, hour, minute)

    def add_person(self, name, date, number, message, eventName):
        row = [name, date, number, message, eventName]
        ws.append(row)
        log1 = open('codeLog.txt', 'r+')
        ir = log1.readline().strip()
        js = int(ir)
        js += 1
        log1.seek(0)
        log1.write(str(js))
        log1.truncate()
        log1.close()
        wb.save('dataM.xlsx')

    def send_messages(self, dt):
        def show_popup_message(head, message):
            popup = CurvedPopup(title=head, content=Label(text=message), size_hint=(None, None), size=(400, 200),
                                separator_height=0, background_color=(1, 1, 1, 1))
            popup.open()

        log2 = open('codeLog.txt', 'r')
        io = log2.readline()
        jo = int(io)
        while jo > 0:
            rows = ws[jo]
            for cell in rows:
                if cell.value == da:
                    number = ws['C' + str(jo)].value
                    message = ws['D' + str(jo)].value
                    h = strftime("%H")
                    hour = int(h)
                    m = strftime("%M")
                    minute = int(m) + 1
                    success = self.send_whatsapp_message(number, message, hour, minute)
                    if success:
                        show_popup_message(head="NOTIFICATION", message="Message Sent")
                    if not success:
                        show_popup_message(head="ERROR", message="Message Not Sent")

            jo = jo - 1
        log2.close()

    def on_start(self):
        Clock.schedule_interval(self.send_messages, 60)

    def show_popup_message(self, head, message):
        popup = CurvedPopup(title=head, content=Label(text=message), size_hint=(None, None), size=(400, 200),
                            separator_height=0, background_color=(1, 1, 1, 1))
        popup.open()

    def search_people(self):
        search_popup = Search(title="FIND PERSON")
        search_popup.open()


if __name__ == '__main__':
    MyApp().run()
